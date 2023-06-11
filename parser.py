from bs4 import BeautifulSoup
from openpyxl import load_workbook
import requests
import fake_useragent

excel_file = 'data.xlsx' #file_name
wb = load_workbook(excel_file) #loading a file
ws = wb['data'] #worksheet in file
ws['A1'] = 'Название'
ws['B1'] = 'Производитель'
ws['C1'] = 'Цена'


link = 'https://aloeapteka.ru'
user = fake_useragent.UserAgent().random
header = {'user-agent':user}
response = requests.get(headers=header, url=link).text
soup = BeautifulSoup(response,'lxml')
block = soup.find('nav',class_='main-nav d-flex justify-content-between')
links = block.find_all('a')

for l in links:
    user = fake_useragent.UserAgent().random
    header = {'user-agent': user}
    href = l.get('href')
    if 'javascript' not in href:
        url = f'{link}{href}'
        response = requests.get(headers=header, url=url).text
        soup = BeautifulSoup(response,'lxml')
        section = soup.find('section',class_='group-catalog')
        if section!=None:
            article = section.find_all('article', class_='collapse-list')
            if len(article)>0:
                for a in article:
                    arcticle_href = a.find_all('a')
                    for ar in arcticle_href:
                        arcticle_href = link + ar.get('href')
                        response = requests.get(headers=header, url=arcticle_href).text
                        soup = BeautifulSoup(response, 'lxml')
                        products_url = soup.find_all('a',class_='img-wr')
                        for p in products_url:
                            product_href = link + p.get('href')
                            products_resp = requests.get(headers=header,url=product_href).text
                            soup_product = BeautifulSoup(products_resp,'lxml')
                            name = soup_product.find('h1').text
                            price = soup_product.find('ins').text
                            producer = soup_product.find('dd').text
                            ws.append([name, price, producer])
                            wb.save(excel_file)
                            print(name)
                        while next != None:
                            soup = BeautifulSoup(response, 'lxml')
                            products_url = soup.find_all('a', class_='img-wr')
                            if soup.find('a', title='Вперед') != None:
                                for p in products_url:
                                    product_href = link + p.get('href')
                                    response_product = requests.get(headers=header, url=product_href).text
                                    soup_product = BeautifulSoup(response_product, 'lxml')
                                    name = soup_product.find('h1').text
                                    price = soup_product.find('ins').text
                                    producer = soup_product.find('dd').text
                                    print(name)
                                    print(product_href)
                                    ws.append([name,price,producer])
                                    wb.save(excel_file)
                                next = soup.find('a', title='Вперед').get('href')
                                next_btn = link + next[22:]
                                print(next_btn)
                                response = requests.get(headers=header, url=next_btn).text
                                soup = BeautifulSoup(response, 'lxml')
                            else:
                                break
            else:
                section_href = soup.find_all('a',class_='btn-collapse')
                for s in section_href:
                    s_href = link + s.get('href')
                    response = requests.get(headers=header, url=s_href).text
                    soup = BeautifulSoup(response, 'lxml')
                    products_url = soup.find_all('a',class_='img-wr')
                    for p in products_url:
                        product_href = link + p.get('href')
                        print(product_href)
                        products_resp = requests.get(headers=header, url=product_href).text
                        soup_product = BeautifulSoup(products_resp, 'lxml')
                        name = soup_product.find('h1').text
                        price = soup_product.find('ins').text
                        producer = soup_product.find('dd').text
                        ws.append([name, price, producer])
                        wb.save(excel_file)
                        print(name)
                    while next!=None:
                        soup = BeautifulSoup(response, 'lxml')
                        products_url = soup.find_all('a', class_='img-wr')
                        if soup.find('a',title='Вперед')!=None:
                                for p in products_url:
                                    product_href = link + p.get('href')
                                    response_product = requests.get(headers=header, url=product_href).text
                                    soup_product = BeautifulSoup(response_product, 'lxml')
                                    name = soup_product.find('h1').text
                                    price = soup_product.find('ins').text
                                    producer = soup_product.find('dd').text
                                    print(name)
                                    ws.append([name,price,producer])
                                    wb.save(excel_file)
                                next = soup.find('a', title='Вперед').get('href')
                                next_btn = link + next[22:]
                                print(next_btn)
                                response = requests.get(headers=header, url=next_btn).text
                                soup = BeautifulSoup(response, 'lxml')
                        else:
                            break